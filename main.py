from __future__ import print_function
import requests
import json
import openpyxl
import re
import os
import string
import math

def load_from_json(file):
    try:
        with open(file, 'r') as myfile:
            return json.load(myfile)
    except IOError:
        with open(file, 'w') as myfile:
            json.dump({}, myfile)
        return {}

config = load_from_json('config.json')
workbook_name = config['workbookName']
email = config['email']
password = config['password']
attributes = config['attributes']
market_attributes = config['marketAttributes']
width = config['width']

def center(text, spacer=' ', length=width, clear=False, display=True):
    if clear:
        os.system('cls' if os.name == 'nt' else 'clear')
    count = int(math.ceil((length - len(text)) / 2))
    if count > 0:
        if display:
            print(spacer * count + text + spacer * count)
        else:
            return (spacer * count + text + spacer * count)
    else:
        if display:
            print(text)
        else:
            return text

class Stockx():
    API_BASE = 'https://stockx.com/api'

    def __init__(self):
        self.customer_id = None
        self.headers = None

    def __api_query(self, request_type, command, data=None):
        endpoint = self.API_BASE + command
        response = None
        if request_type == 'GET':
            response = requests.get(endpoint, params=data, headers=self.headers)
        elif request_type == 'POST':
            response = requests.post(endpoint, json=data, headers=self.headers)
        elif request_type == 'DELETE':
            response = requests.delete(endpoint, json=data, headers=self.headers)
        return response.json()

    def __get(self, command, data=None):
        return self.__api_query('GET', command, data)

    def __post(self, command, data=None):
        return self.__api_query('POST', command, data)

    def __delete(self, command, data=None):
        return self.__api_query('DELETE', command, data)

    def authenticate(self, email, password):
        endpoint = self.API_BASE + '/login'
        payload = {
            'email': email,
            'password': password
        }
        response = requests.post(endpoint, json=payload)
        customer = response.json().get('Customer', None)
        if customer is None:
            raise ValueError('Authentication failed, check username/password')
        self.customer_id = response.json()['Customer']['id']
        self.headers = {
            'JWT-Authorization': response.headers['jwt-authorization']
        }
        return True

    def selling(self):
        command = '/customers/{0}/selling'.format(self.customer_id)
        response = self.__get(command)
        return response['PortfolioItems']

stockx = Stockx()

def json_to_title(json):
	return re.sub(r'(\w)([A-Z])', r'\1 \2', json).title().replace('Shoe Size', 'Size')

def setup_workbook(attributes, market_attributes, workbook_name):
	wb = openpyxl.Workbook()
	ws = wb.active
	headers = [json_to_title(item) for item in attributes + market_attributes]
	for i in range(0, len(headers)):
		cell = ws[list(string.ascii_uppercase)[i + list(string.ascii_uppercase).index('B')] + str(2)]
		cell.value = headers[i]
		cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
	wb.save(workbook_name)

def write_workbook(email, password, attributes, market_attributes, workbook_name):
	wb = openpyxl.load_workbook(workbook_name)
	ws = wb.active
	if stockx.authenticate(email, password):
		i = 0
		for item in stockx.selling():
			if item['text'] == 'Asking':
				product = item['product']
				for k in range(0, len(attributes)):
					cell = ws[list(string.ascii_uppercase)[k + list(string.ascii_uppercase).index('B')] + str(3 + i)]
					cell.value = product[attributes[k]]
					cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
				if market_attributes != []:
					market = product['market']
					for k in range(len(attributes), len(attributes) + len(market_attributes)):
						cell = ws[list(string.ascii_uppercase)[k + list(string.ascii_uppercase).index('B')] + str(3 + i)]
						cell.value = market[market_attributes[k - len(attributes)]]
						cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
				i += 1
	wb.save(workbook_name)

center(' ', clear=True)
center('StockX to Excel by @DefNotAvg')
center('-', '-')
print('{}\r'.format(center('Setting up the Excel Workbook...', display=False)), end='')
setup_workbook(attributes, market_attributes, workbook_name)
center('Successfully set up the Excel Workbook!!')
print('{}\r'.format(center('Writing StockX data to {}...'.format(workbook_name), display=False)), end='')
write_workbook(email, password, attributes, market_attributes, workbook_name)
center('Successfully exported StockX data to {}!!'.format(workbook_name))